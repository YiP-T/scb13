#!/usr/bin/env python 
# -*- coding: utf-8 -*- 
# @Time : 2020/5/21 8:22
# @Author : Lemon_YiP
# @QQ: 693625157


'''
1、用例，读取测试数据 ---Done == read_data()
2、用数据发送接口请求，执行测试 --- Done == post_func()
3、执行结果 vs 预期结果 == 得出测试结果  ----用断言做对比
4、得到的测试结果，回写到测试用例表格里 --- Done == write_result()
5、将整个测试过程封装成一个执行函数 --- Done == execute_func()

'''
import openpyxl
import requests

session = requests.session()  # requests库的session模块，定义一个变量


# 读取测试用例的函数
def read_data(filename, sheetname):
    wb = openpyxl.load_workbook(filename)  # 加入这个工作簿对象 --赋值给变量wb
    sheet = wb[sheetname]  # 表单
    max_row = sheet.max_row  # 获取最大行号
    cases = []  # 准备一个空列表，存放数据
    for i in range(2, max_row + 1):
        case = dict(
            case_id=sheet.cell(row=i, column=1).value,  # 获取编号
            url=sheet.cell(row=i, column=5).value,  # 获取upl
            data=sheet.cell(row=i, column=6).value,  # 获取data
            expected_result=sheet.cell(row=i, column=7).value  # 获取期望结果
        )  # 一个用例放入一个字典里
        cases.append(case)  # 把字典追加到列表里保存起来，方便后续取值应用
    # print(cases)
    return cases  # 定义返回值


# 发送接口请求的函数
def post_func(qcd_url, qcd_data):
    res = session.post(url=qcd_url, data=qcd_data)  # post方法发送接口请求
    result = res.json()  # 用json获取响应结果 -- 字典
    return result  # 定义返回值


# 写入测试结果的函数
def write_result(filename, sheetname, row, column, real_result):
    wb = openpyxl.load_workbook(filename)  # 加载工作薄 --赋值给一个变量
    sheet = wb[sheetname]
    sheet.cell(row=row, column=column).value = real_result  # 写入
    wb.save(filename)


# 将整个测试过程封装成一个执行函数
def execute_func(filename, sheetname):
    test_cases = read_data(filename, sheetname)  # 调用读取数据的函数
    for test_case in test_cases:
        case_id = test_case.get('case_id')  # 获取对应的case_id
        url = test_case.get('url')  # 获取对应的url
        data = test_case.get('data')  # 获取对应的data
        data = eval(data)  # eval（）进行数据类型转化 -- 字符串 -->字典
        expected_result = test_case.get('expected_result')  # 获取对应的expected_result
        expected_result = expected_result.replace('null', 'None')
        # python不认识null，用replace函数进行文本替换 -- null -->None
        expected_result = eval(expected_result)
        # print(type(expected_result))
        # 期望结果和执行结果进行对比，需要统一数据类型，用eval函数转换 -- 字符串 -->字典
        # print(case_id, url, data, expected_result)
        real_result = post_func(qcd_url=url, qcd_data=data)  # 调用发送接口请求的函数
        real_msg = real_result.get('msg')  # 字典取值 -- 获取做断言的有效数字段
        expected_msg = expected_result.get('msg')
        print('期望测试结果是：{}'.format(expected_msg))
        print('真是执行结果是：{}'.format(real_msg))
        if expected_msg == real_msg:
            print('第{}条测试用例通过：'.format(case_id))
            final_result = 'passed'
        else:
            print('第{}条测试用例不通过：'.format(case_id))
            final_result = 'failed'
        print('**' * 15)
        write_result(filename, sheetname, case_id + 1, 8, final_result)
        # 调用写入测试结果的函数 -- 写入


execute_func('test_case.xlsx', 'register')
# 调用执行函数
